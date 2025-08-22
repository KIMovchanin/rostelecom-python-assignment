from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QVBoxLayout, QHBoxLayout, QComboBox, QTextEdit
)
from PyQt6.QtCore import Qt
import sys
from openpyxl import load_workbook


class ExcelFilterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Фильтр Excel — PyQt6 + openpyxl (Python 3.9)")
        self.resize(800, 420)
        self.header_row = None
        self.data_start_row = None
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.input_edit = QLineEdit()
        self.input_btn = QPushButton("Выбрать файл…")
        input_row = QHBoxLayout()
        input_row.addWidget(QLabel("Входной .xlsx:"))
        input_row.addWidget(self.input_edit)
        input_row.addWidget(self.input_btn)

        self.column_combo = QComboBox()
        self.value_edit = QLineEdit()
        col_row = QHBoxLayout()
        col_row.addWidget(QLabel("Столбец:"))
        col_row.addWidget(self.column_combo, 1)
        col_row.addWidget(QLabel("Значение:"))
        col_row.addWidget(self.value_edit, 1)

        self.output_edit = QLineEdit()
        self.output_btn = QPushButton("Сохранить как…")
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel("Результат:"))
        output_row.addWidget(self.output_edit)
        output_row.addWidget(self.output_btn)

        self.run_btn = QPushButton("Выполнить фильтрацию")
        self.log = QTextEdit()
        self.log.setReadOnly(True)

        layout = QVBoxLayout()
        layout.addLayout(input_row)
        layout.addLayout(col_row)
        layout.addLayout(output_row)
        layout.addWidget(self.run_btn)
        layout.addWidget(QLabel("Лог:"))
        layout.addWidget(self.log, 1)
        self.setLayout(layout)

    def _connect_signals(self):
        self.input_btn.clicked.connect(self.browse_input)
        self.output_btn.clicked.connect(self.browse_output)
        self.run_btn.clicked.connect(self.run_filter)

    def browse_input(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel-файл", "", "Excel (*.xlsx)")
        if path:
            self.input_edit.setText(path)
            self.populate_columns(Path(path))

    def browse_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить результат", "", "Excel (*.xlsx)")
        if path and not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        if path:
            self.output_edit.setText(path)

    def populate_columns(self, path: Path):
        self.column_combo.clear()
        self.header_row = None
        self.data_start_row = None
        try:
            # открываем книгу на чтение.
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            # берём активный лист
            ws = wb.active

            header_row = self._guess_header_row(ws)

            row_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            headers = [str(cell_value).strip() for cell_value in row_vals if cell_value not in (None, "")]

            if not headers:
                self.column_combo.addItems(["Заголовки не найдены."])
                self.log.append("Не удалось распознать строку заголовков.")
                return

            self.column_combo.addItems(headers)
            self.header_row = header_row
            self.data_start_row = header_row + 1
            self.log.append(f"OK: Загружены столбцы из строки {header_row}: {', '.join(headers)}")
        except FileNotFoundError:
            self.log.append("Файл не найден. Проверь путь.")
        except Exception as e:
            self.log.append(f"Ошибка при чтении Excel: {e!r}")

    @staticmethod
    def _norm(v) -> str:
        return str(v).strip().casefold() if v is not None else ""

    def _guess_header_row(self, ws, search_limit: int = 25) -> int:
        expected = {self._norm(header_name) for header_name in ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]}
        best_row = 1
        best_score = -1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=search_limit, values_only=True), start=1):
            vals = [cell for cell in row if cell not in (None, "")]
            non_empty = len(vals)
            if non_empty == 0:
                continue
            lower_set = {self._norm(c) for c in vals}
            score = non_empty
            score += 2 * len(lower_set & expected) # чем больше совпадений с ожидаемым результатом - тогда берём
            if score > best_score:
                best_score = score
                best_row = idx
        return best_row

    def run_filter(self):
        self.log.append("Нажата кнопка \"Выполнить фильтрацию\" — логику добавим на следующем шаге.")


def main():
    app = QApplication(sys.argv)
    win = ExcelFilterApp()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
