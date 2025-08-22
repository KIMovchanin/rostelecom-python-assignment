from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QVBoxLayout, QHBoxLayout, QComboBox, QTextEdit
)
from PyQt6.QtCore import Qt
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date

class ExcelFilterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Фильтр Excel")
        self.resize(800, 420)
        self.header_row = None
        self.data_start_row = None
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.input_edit = QLineEdit()
        self.input_btn = QPushButton("Выбрать файл…")
        input_row = QHBoxLayout()
        input_row.addWidget(QLabel("Входной .xlsx:"))
        input_row.addWidget(self.input_edit)
        input_row.addWidget(self.input_btn)

        self.column_combo = QComboBox()
        self.value_edit = QLineEdit()
        col_row = QHBoxLayout()
        col_row.addWidget(QLabel("Столбец:"))
        col_row.addWidget(self.column_combo, 1)
        col_row.addWidget(QLabel("Значение:"))
        col_row.addWidget(self.value_edit, 1)

        self.output_edit = QLineEdit()
        self.output_btn = QPushButton("Сохранить как…")
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel("Результат:"))
        output_row.addWidget(self.output_edit)
        output_row.addWidget(self.output_btn)

        self.run_btn = QPushButton("Выполнить фильтрацию")
        self.log = QTextEdit()
        self.log.setReadOnly(True)

        layout = QVBoxLayout()
        layout.addLayout(input_row)
        layout.addLayout(col_row)
        layout.addLayout(output_row)
        layout.addWidget(self.run_btn)
        layout.addWidget(QLabel("Лог:"))
        layout.addWidget(self.log, 1)
        self.setLayout(layout)

    def _connect_signals(self):
        self.input_btn.clicked.connect(self.browse_input)
        self.output_btn.clicked.connect(self.browse_output)
        self.run_btn.clicked.connect(self.run_filter)

    def browse_input(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel-файл", "", "Excel (*.xlsx)")
        if path:
            self.input_edit.setText(path)
            self.populate_columns(Path(path))

    def browse_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить результат", "", "Excel (*.xlsx)")
        if path and not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        if path:
            self.output_edit.setText(path)

    def populate_columns(self, path: Path):
        self.column_combo.clear()
        self.header_row = None
        self.data_start_row = None
        try:
            # открываем книгу на чтение.
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            # берём активный лист
            ws = wb.active

            header_row = self._guess_header_row(ws)

            row_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            headers = [str(cell_value).strip() for cell_value in row_vals if cell_value not in (None, "")]

            if not headers:
                self.column_combo.addItems(["Заголовки не найдены."])
                self.log.append("Не удалось распознать строку заголовков.")
                return

            self.column_combo.addItems(headers)
            self.header_row = header_row
            self.data_start_row = header_row + 1
            self.log.append(f"OK: Загружены столбцы из строки {header_row}: {', '.join(headers)}")
        except FileNotFoundError:
            self.log.append("Файл не найден. Проверь путь.")
        except Exception as e:
            self.log.append(f"Ошибка при чтении Excel: {e!r}")

    @staticmethod
    def _norm(v) -> str:
        return str(v).strip().casefold() if v is not None else ""

    def _norm_cell_for_compare(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, str):
            d = self._try_parse_date(v)
            if d:
                return d.isoformat()
        return str(v).strip().casefold()

    def _try_parse_date(self, s: str):
        s = s.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _guess_header_row(self, ws, search_limit: int = 25) -> int:
        expected = {self._norm(header_name) for header_name in ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]}
        best_row = 1
        best_score = -1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=search_limit, values_only=True), start=1):
            vals = [cell for cell in row if cell not in (None, "")]
            non_empty = len(vals)
            if non_empty == 0:
                continue
            lower_set = {self._norm(c) for c in vals}
            score = non_empty
            score += 2 * len(lower_set & expected) # чем больше совпадений с ожидаемым результатом - тогда берём
            if score > best_score:
                best_score = score
                best_row = idx
        return best_row

    REQUIRED_OUT_HEADERS = ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]

    def _make_header_map(self, ws, header_row: int):
        row_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers = [str(v).strip() if v not in (None, "") else "" for v in row_vals]

        norm_to_index = {}
        norm_to_original = {}
        for idx, name in enumerate(headers):
            if not name:
                continue
            key = self._norm(name)
            if key not in norm_to_index:
                norm_to_index[key] = idx # создаём ключ - норм. название столбца, значение - его индекс в списке
                norm_to_original[key] = name # ключ - нормализованное название, значение - название как в Excel
        return norm_to_index, norm_to_original

    def run_filter(self):
        in_path = self.input_edit.text().strip()
        out_path = self.output_edit.text().strip()
        col_display = self.column_combo.currentText().strip()
        value_raw = self.value_edit.text().strip()
        value_date = self._try_parse_date(value_raw)

        if not in_path:
            self.log.append("Не выбран входной файл.")
            return
        if not out_path:
            self.log.append("Не выбран путь для сохранения результата.")
            return
        if not col_display:
            self.log.append("Не выбран столбец для фильтра.")
            return
        if self.header_row is None or self.data_start_row is None:
            self.log.append("Не определена строка заголовков. Выберите файл заново.")
            return

        try:
            wb_in = load_workbook(filename=in_path, read_only=True, data_only=True)
            ws_in = wb_in.active
            norm_to_index, _ = self._make_header_map(ws_in, self.header_row)

            filter_key = self._norm(col_display)
            filter_col_idx = norm_to_index.get(filter_key)
            if filter_col_idx is None:
                self.log.append(f"В файле не найден столбец: {col_display}")
                return

            wanted = []
            missing = []
            for name in self.REQUIRED_OUT_HEADERS:
                idx = norm_to_index.get(self._norm(name))
                if idx is None:
                    missing.append(name)
                else:
                    wanted.append((name, idx))

            if not wanted:
                self.log.append("Не найдено ни одной из колонок: " + ", ".join(self.REQUIRED_OUT_HEADERS))
                return
            if missing:
                self.log.append("Отсутствуют колонки: " + ", ".join(missing))

            value_date = self._try_parse_date(value_raw)
            if value_date:
                value_norm = value_date.isoformat()
            else:
                value_norm = self._norm(value_raw)

            matched_rows = []
            for row in ws_in.iter_rows(min_row=self.data_start_row, values_only=True):
                cell = row[filter_col_idx] if filter_col_idx < len(row) else None
                cell_norm = self._norm_cell_for_compare(cell)

                if cell_norm == value_norm:
                    out_values = []
                    for name, idx in wanted:
                        row_value = row[idx] if idx < len(row) else None
                        if name == "Дата найма" and isinstance(row_value, str):
                            d = self._try_parse_date(row_value)
                            if d:
                                row_value = d
                        out_values.append(row_value)
                    matched_rows.append(out_values)

            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "Результат"

            header_names = [name for name, _ in wanted]
            ws_out.append(header_names)

            # данные
            for out_row in matched_rows:
                ws_out.append(out_row)

            if "Дата найма" in header_names:
                col_idx = header_names.index("Дата найма") + 1
                col_letter = get_column_letter(col_idx)
                for cell in ws_out[col_letter][1:]:
                    if isinstance(cell.value, (datetime, date)):
                        cell.number_format = "DD.MM.YYYY"
            try:
                wb_out.save(out_path)
                count = len(matched_rows)
                if count == 0:
                    self.log.append(f"Под критерий ничего не подошло. Создан файл только с заголовком: {out_path}")
                else:
                    self.log.append(f"Сохранено {count} строк в файл: {out_path}")
            except PermissionError:
                self.log.append("Не удалось сохранить: файл открыт в Excel/LibreOffice. Закройте его и повторите.")

        except Exception as e:
            self.log.append(f"Ошибка фильтрации/сохранения: {e!r}")


def main():
    app = QApplication(sys.argv)
    win = ExcelFilterApp()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
